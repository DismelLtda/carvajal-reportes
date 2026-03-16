"""
📊 PARSER BASE DE EXCEL
Clase abstracta para parsear reportes Excel (VENTAS e INVENTARIO)
Encapsula lógica común: lectura, validación, extracción de metadata
"""

from abc import ABC, abstractmethod
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from dataclasses import dataclass
from datetime import datetime
import hashlib
import logging
from typing import Any, List, Dict, Optional

logger = logging.getLogger(__name__)


@dataclass
class Metadata:
    """Metadatos de un reporte Excel"""
    archivo: str
    tipo_reporte: str
    fecha_descarga: datetime
    tamaño_bytes: int
    hash_contenido: str
    numero_filas: int
    numero_columnas: int
    fila_cabecera: int
    fila_inicio_datos: int
    fecha_reporte: Optional[datetime] = None  # Fecha real del reporte extraída del Excel
    id_informe: str = ""
    proveedor_emisor: str = ""
    entidad_receptor: str = ""


@dataclass  
class FilaDetalle:
    """Una fila de detalle del reporte"""
    numero_fila: int
    datos: Dict[str, Any]
    valores_vacios: List[str]  # Columnas vacías


class ExcelParserBase(ABC):
    """
    Clase base para parsers de reportes Excel
    
    Subclases implementan lógica específica de tipo de reporte:
    - ExcelParserVentas: 14 columnas, datos transaccionales
    - ExcelParserInventario: 13 columnas, datos de stock
    """
    
    # Configuración por defecto (sobreescribir en subclases)
    FILA_CABECERA = 9
    FILA_INICIO_DATOS = 10
    NUMERO_COLUMNAS_ESPERADO = 14
    COLUMNAS = []  # Lista de nombres de columnas esperadas
    
    def __init__(self, ruta_archivo: Path):
        """
        Inicializa parser
        
        Args:
            ruta_archivo: Path del archivo Excel a procesar
        """
        self.ruta_archivo = Path(ruta_archivo)
        self.workbook = None
        self.worksheet = None
        self.metadata: Optional[Metadata] = None
        self._contenido_para_hash = []
        
        # Validar archivo
        self._validar_archivo()
    
    def _validar_archivo(self):
        """Valida que el archivo existe y tiene extensión correcta"""
        if not self.ruta_archivo.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {self.ruta_archivo}")
        
        if self.ruta_archivo.suffix.lower() not in {'.xlsx', '.xls'}:
            raise ValueError(f"Formato inválido: {self.ruta_archivo.suffix}")
        
        logger.debug(f"✅ Archivo validado: {self.ruta_archivo.name}")
    
    def abrir(self):
        """Abre el archivo Excel y lo prepara para lectura"""
        try:
            logger.debug(f"Abriendo {self.ruta_archivo}...")
            self.workbook = load_workbook(self.ruta_archivo, data_only=True)
            self.worksheet = self.workbook.active
            logger.info(f"✅ Archivo abierto: {self.ruta_archivo.name}")
        except Exception as e:
            logger.error(f"❌ Error abriendo archivo: {str(e)}")
            raise
    
    def cerrar(self):
        """Cierra el archivo Excel"""
        if self.workbook:
            self.workbook.close()
            logger.debug("✅ Archivo cerrado")
    
    def extraer_metadata(self) -> Metadata:
        """
        Extrae metadatos del archivo Excel
        """
        
        if not self.worksheet:
            raise ValueError("Archivo no abierto. Llama a .abrir() primero")

        # Contar filas de datos
        numero_filas = self._contar_filas_datos()
        
        # Contar columnas
        numero_columnas = self._contar_columnas()

        # Extraer ID del informe de A1
        # Ejemplo: "Reporte de Inventario: 80008987202026030616231"
        celda_a1 = str(self.worksheet['A1'].value or '')
        id_informe = ""
        if ':' in celda_a1:
            id_informe = celda_a1.split(':', 1)[1].strip()
        
        # Extraer emisor/proveedor de Fila 3, Col 2
        valor_f3 = str(self.worksheet.cell(row=3, column=2).value or '').strip()
        
        # Extraer receptor/entidad de Fila 4, Col 2
        valor_f4 = str(self.worksheet.cell(row=4, column=2).value or '').strip()

        # Extraer fecha real del reporte (usualmente fila 2 o 5)
        # Intentamos fila 2, columna 2 primero
        valor_fecha = self.worksheet.cell(row=2, column=2).value
        fecha_real = None
        
        if isinstance(valor_fecha, datetime):
            fecha_real = valor_fecha
        elif isinstance(valor_fecha, str):
            # Intentar parsear si es string (ej: "10/03/2026")
            try:
                # Carvajal suele usar DD/MM/YYYY o DD-MM-YYYY
                from dateutil import parser as date_parser
                fecha_real = date_parser.parse(valor_fecha, dayfirst=True)
            except:
                logger.warning(f"No se pudo parsear fecha en fila 2: {valor_fecha}")
        
        # Si no hay fecha en fila 2, intentamos fila 5 (algunos formatos)
        if not fecha_real:
            valor_f5 = self.worksheet.cell(row=5, column=2).value
            if isinstance(valor_f5, datetime):
                fecha_real = valor_f5
            elif isinstance(valor_f5, str):
                try:
                    from dateutil import parser as date_parser
                    fecha_real = date_parser.parse(valor_f5, dayfirst=True)
                except:
                    pass

        self.metadata = Metadata(
            archivo=self.ruta_archivo.name,
            tipo_reporte=self.__class__.__name__,
            fecha_descarga=datetime.now(),
            fecha_reporte=fecha_real,
            tamaño_bytes=self.ruta_archivo.stat().st_size,
            hash_contenido=self._calcular_hash(),
            numero_filas=numero_filas,
            numero_columnas=numero_columnas,
            fila_cabecera=self.FILA_CABECERA,
            fila_inicio_datos=self.FILA_INICIO_DATOS,
            id_informe=id_informe,
            proveedor_emisor=valor_f3,
            entidad_receptor=valor_f4
        )
        
        logger.info(
            f"📊 Metadata extraída:\n"
            f"   Filas: {numero_filas}\n"
            f"   Columnas: {numero_columnas}\n"
            f"   Hash: {self.metadata.hash_contenido[:16]}...\n"
            f"   Tamaño: {self.metadata.tamaño_bytes / 1024:.1f} KB"
        )
        
        return self.metadata
    
    def _calcular_hash(self) -> str:
        """
        Calcula SHA256 del contenido del archivo
        Usado para detectar duplicados
        
        Returns:
            Hash SHA256 en hexadecimal
        """
        
        sha256 = hashlib.sha256()
        
        with open(self.ruta_archivo, 'rb') as f:
            # Leer en chunks de 64KB
            for chunk in iter(lambda: f.read(65536), b''):
                sha256.update(chunk)
        
        hash_hex = sha256.hexdigest()
        logger.debug(f"Hash calculado: {hash_hex[:16]}...")
        
        return hash_hex
    
    def _contar_filas_datos(self) -> int:
        """
        Cuenta el número total de filas con datos
        Desde FILA_INICIO_DATOS hasta última fila con datos en columna A
        
        Returns:
            Número de filas
        """
        
        contador = 0
        
        for fila_num in range(self.FILA_INICIO_DATOS, self.worksheet.max_row + 1):
            celda = self.worksheet.cell(row=fila_num, column=1)
            
            if celda.value is not None:
                contador += 1
            elif contador > 0:
                # Si encontramos vacía después de datos, paramos
                break
        
        return contador
    
    def _contar_columnas(self) -> int:
        """
        Cuenta el número de columnas con encabezados
        Analiza FILA_CABECERA desde primera hasta última con datos
        
        Returns:
            Número de columnas
        """
        
        contador = 0
        
        for col_num in range(1, 50):
            celda = self.worksheet.cell(row=self.FILA_CABECERA, column=col_num)
            
            if celda.value is not None:
                contador += 1
        
        return contador
    
    def extraer_cabeceras(self) -> List[str]:
        """
        Extrae los nombres de columnas de FILA_CABECERA
        
        Returns:
            Lista de nombres de columnas
        """
        
        cabeceras = []
        # Leer hasta el número esperado o un máximo de 50
        max_cols = max(50, self.NUMERO_COLUMNAS_ESPERADO + 1)
        
        for col_num in range(1, max_cols):
            celda = self.worksheet.cell(row=self.FILA_CABECERA, column=col_num)
            
            # Si llegamos al final de los datos del Excel y no hay más columnas, paramos
            if col_num > self.worksheet.max_column and celda.value is None:
                break
                
            val = str(celda.value).strip() if celda.value is not None else f"COL_{col_num}"
            
            # Manejar duplicados
            base_val = val
            counter = 1
            while val in cabeceras:
                val = f"{base_val}_{counter}"
                counter += 1
                
            cabeceras.append(val)
        
        # Recortar cabeceras vacías al final si exceden el esperado
        while len(cabeceras) > self.NUMERO_COLUMNAS_ESPERADO and cabeceras[-1].startswith("COL_"):
            cabeceras.pop()
            
        logger.debug(f"Cabeceras extraídas: {cabeceras}")
        
        return cabeceras
    
    def extraer_detalles(self) -> List[FilaDetalle]:
        """
        Extrae todas las filas de detalle del reporte
        
        Returns:
            Lista de FilaDetalle con datos y validación
        """
        
        detalles = []
        cabeceras = self.extraer_cabeceras()
        
        logger.info(f"Extrayendo detalles desde fila {self.FILA_INICIO_DATOS}...")
        
        for fila_num in range(self.FILA_INICIO_DATOS, self.worksheet.max_row + 1):
            # Extraer datos de la fila
            datos = {}
            valores_vacios = []
            
            for col_idx, cabecera in enumerate(cabeceras, start=1):
                celda = self.worksheet.cell(row=fila_num, column=col_idx)
                valor = celda.value
                
                if valor is None:
                    valores_vacios.append(cabecera)
                else:
                    # Limpiar espacios en blanco si es texto para evitar inconsistencias
                    if isinstance(valor, str):
                        valor = valor.strip()
                    datos[cabecera] = valor
            
            # Validar mínimo de datos (al menos una columna con valor)
            if not datos:
                # Fila completamente vacía, fin de datos
                break
            
            fila_detalle = FilaDetalle(
                numero_fila=fila_num,
                datos=datos,
                valores_vacios=valores_vacios
            )
            
            detalles.append(fila_detalle)
        
        logger.info(f"✅ {len(detalles)} filas de detalle extraídas")
        
        return detalles
    
    def validar_estructura(self) -> tuple[bool, List[str]]:
        """
        Valida que la estructura del reporte sea correcta
        
        Returns:
            Tupla (es_válido, lista_errores)
        """
        
        errores = []
        
        # Validar cabecera existe
        cabecera_valor = self.worksheet.cell(row=self.FILA_CABECERA, column=1).value
        if not cabecera_valor:
            errores.append(f"Fila de cabecera {self.FILA_CABECERA} vacía")
        
        # Validar datos
        datos_valor = self.worksheet.cell(row=self.FILA_INICIO_DATOS, column=1).value
        if not datos_valor:
            errores.append(f"No hay datos en fila {self.FILA_INICIO_DATOS}")
        
        # Validar número de columnas
        cabeceras = self.extraer_cabeceras()
        num_cols = len(cabeceras)
        
        if num_cols != self.NUMERO_COLUMNAS_ESPERADO:
            logger.warning(f"⚠️  Diferencia de columnas en {self.ruta_archivo.name}:")
            logger.warning(f"   Esperadas: {self.NUMERO_COLUMNAS_ESPERADO}, Encontradas: {num_cols}")
            logger.warning(f"   Cabeceras encontradas: {cabeceras}")
            
            # Si hay al menos un mínimo de columnas, permitimos continuar
            if num_cols < (self.NUMERO_COLUMNAS_ESPERADO - 5):
                errores.append(f"Demasiadas pocas columnas encontradas ({num_cols})")
        
        # Validar mínimo de filas
        num_filas = self._contar_filas_datos()
        if num_filas == 0:
            errores.append("No hay filas de datos")
        
        es_valido = len(errores) == 0
        
        if es_valido:
            logger.info("✅ Estructura válida")
        else:
            logger.warning(f"⚠️  Estructura inválida: {errores}")
        
        return es_valido, errores
    
    @abstractmethod
    def procesar(self) -> Dict[str, Any]:
        """
        Procesa el archivo completo
        Implementar en subclases específicas
        
        Returns:
            Diccionario con resultados del procesamiento
        """
        pass
    
    def __enter__(self):
        """Context manager: abrir archivo"""
        self.abrir()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager: cerrar archivo"""
        self.cerrar()


# ============================================================================
# EJEMPLO DE USO
# ============================================================================

if __name__ == '__main__':
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Este módulo es base, no se usa directamente
    print("Este es un módulo base - ver ExcelParserVentas y ExcelParserInventario")
