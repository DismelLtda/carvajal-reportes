"""
🔍 DETECTOR DE TIPOS DE REPORTES
Identifica automáticamente si un archivo Excel es VENTAS o INVENTARIO
mediante análisis de contenido (primera celda del rango de datos)
"""

from pathlib import Path
from openpyxl import load_workbook
from enum import Enum
import logging

logger = logging.getLogger(__name__)


class TipoReporte(Enum):
    """Tipos de reportes soportados"""
    VENTAS = 'VENTAS'
    INVENTARIO = 'INVENTARIO'
    DESCONOCIDO = 'DESCONOCIDO'


class ReporteDetector:
    """
    Detecta automáticamente el tipo de reporte Excel analizando:
    1. Primera fila (A1): "Reporte de Ventas" vs "Reporte de Inventario"
    2. Estructura de columnas
    3. Número de filas de datos
    
    Estrategia:
    - VENTAS: 60-100 filas de datos, 14 columnas, encabezado "Reporte de Ventas"
    - INVENTARIO: 8000-20000 filas de datos, 13 columnas, encabezado "Reporte de Inventario"
    """
    
    # Configuración de detección
    PATRONES = {
        TipoReporte.INVENTARIO: {
            'palabras_clave': ['reporte de inventario', 'inventario', 'inventory', 'stock'],
            'columnas_esperadas': 12, # Ajustado a realidad (incluye columnas vacías)
            'filas_esperadas_min': 5, # Bajamos el mínimo para reportes pequeños
            'filas_esperadas_max': 50000,
            'fila_cabecera': 9,
            'fila_inicio_datos': 10,
        },
        TipoReporte.VENTAS: {
            'palabras_clave': ['reporte de ventas', 'ventas', 'sales'],
            'columnas_esperadas': 13, # Ajustado a realidad
            'filas_esperadas_min': 5, # Bajamos el mínimo para reportes pequeños
            'filas_esperadas_max': 5000,
            'fila_cabecera': 9,
            'fila_inicio_datos': 10,
        }
    }
    
    @staticmethod
    def detectar_tipo(ruta_archivo: Path) -> TipoReporte:
        """
        Detecta el tipo de reporte analizando el contenido del archivo Excel
        """
        ruta_archivo = Path(ruta_archivo)
        
        # Estrategia 0: Analizar nombre de archivo (Fallback rápido)
        nombre = ruta_archivo.name.lower()
        if 'inventario' in nombre:
            logger.info(f"✅ Tipo detectado por nombre: INVENTARIO")
            return TipoReporte.INVENTARIO
        if 'venta' in nombre:
            logger.info(f"✅ Tipo detectado por nombre: VENTAS")
            return TipoReporte.VENTAS
        
        # Validar existencia
        if not ruta_archivo.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta_archivo}")
        
        # Validar extensión
        if ruta_archivo.suffix.lower() not in {'.xlsx', '.xls'}:
            raise ValueError(f"Formato no soportado: {ruta_archivo.suffix}")
        
        try:
            # Cargar libro Excel
            wb = load_workbook(ruta_archivo, data_only=True)
            ws = wb.active
            
            # Estrategia 1: Analizar celda A1 (encabezado principal)
            encabezado = str(ws['A1'].value or '').lower().strip()
            logger.debug(f"Encabezado (A1): '{encabezado}'")
            
            tipo_detectado = ReporteDetector._detectar_por_encabezado(encabezado)
            if tipo_detectado != TipoReporte.DESCONOCIDO:
                logger.info(f"✅ Tipo detectado por encabezado: {tipo_detectado.value}")
                return tipo_detectado
            
            # Estrategia 2: Analizar número de columnas y filas
            tipo_detectado = ReporteDetector._detectar_por_estructura(ws)
            if tipo_detectado != TipoReporte.DESCONOCIDO:
                logger.info(f"✅ Tipo detectado por estructura: {tipo_detectado.value}")
                return tipo_detectado
            
            logger.warning(f"⚠️  No se pudo determinar tipo: {ruta_archivo.name}")
            return TipoReporte.DESCONOCIDO
            
        except Exception as e:
            logger.error(f"❌ Error al detectar tipo: {str(e)}")
            raise
        finally:
            if 'wb' in locals():
                wb.close()
    
    @staticmethod
    def _detectar_por_encabezado(encabezado: str) -> TipoReporte:
        """
        Detecta tipo analizando el encabezado (A1)
        
        Args:
            encabezado: Contenido de la celda A1 en minúsculas
            
        Returns:
            TipoReporte
        """
        
        for tipo, config in ReporteDetector.PATRONES.items():
            for palabra in config['palabras_clave']:
                if palabra in encabezado:
                    return tipo
        
        return TipoReporte.DESCONOCIDO
    
    @staticmethod
    def _detectar_por_estructura(ws) -> TipoReporte:
        """
        Detecta tipo analizando estructura (columnas y filas)
        
        Args:
            ws: Hoja de Excel (openpyxl Worksheet)
            
        Returns:
            TipoReporte
        """
        
        try:
            # Contar columnas con datos en fila 9 (cabecera)
            fila_cabecera = 9
            columnas_con_datos = 0
            
            for col in range(1, 30):  # Verificar hasta 30 columnas
                celda = ws.cell(row=fila_cabecera, column=col)
                if celda.value is not None:
                    columnas_con_datos += 1
            
            # Contar filas de datos (desde fila 10 hasta última con datos)
            fila_inicio = 10
            filas_con_datos = 0
            
            for fila in range(fila_inicio, ws.max_row + 1):
                # Verificar si fila tiene datos en primera columna
                if ws.cell(row=fila, column=1).value is not None:
                    filas_con_datos += 1
            
            logger.debug(f"Estructura: {columnas_con_datos} columnas, {filas_con_datos} filas")
            
            # Comparar con patrones esperados
            for tipo, config in ReporteDetector.PATRONES.items():
                cols_esperadas = config['columnas_esperadas']
                filas_min = config['filas_esperadas_min']
                filas_max = config['filas_esperadas_max']
                
                # Tolerancia: ±2 columnas, ±10% de filas
                cols_ok = abs(columnas_con_datos - cols_esperadas) <= 2
                filas_ok = filas_min <= filas_con_datos <= filas_max
                
                if cols_ok and filas_ok:
                    return tipo
            
            return TipoReporte.DESCONOCIDO
            
        except Exception as e:
            logger.error(f"Error analizando estructura: {str(e)}")
            return TipoReporte.DESCONOCIDO
    
    @staticmethod
    def validar_deteccion(ruta_archivo: Path) -> tuple[TipoReporte, dict]:
        """
        Valida la detección retornando tipo + metadata
        
        Args:
            ruta_archivo: Ruta del archivo Excel
            
        Returns:
            Tupla (tipo_reporte, metadata_dict)
            
        Metadata incluye:
            - 'tipo': TipoReporte
            - 'archivo': nombre
            - 'tamaño_mb': tamaño en MB
            - 'confianza': porcentaje de confianza
            - 'detalles': información adicional
        """
        
        ruta_archivo = Path(ruta_archivo)
        
        tipo = ReporteDetector.detectar_tipo(ruta_archivo)
        
        # Calcular metadatos
        tamaño_bytes = ruta_archivo.stat().st_size
        tamaño_mb = tamaño_bytes / (1024 * 1024)
        
        metadata = {
            'tipo': tipo.value,
            'archivo': ruta_archivo.name,
            'tamaño_mb': round(tamaño_mb, 2),
            'confianza': 'ALTA' if tipo != TipoReporte.DESCONOCIDO else 'BAJA',
            'ruta_absoluta': str(ruta_archivo),
        }
        
        return tipo, metadata


# ============================================================================
# UTILIDADES
# ============================================================================

def detectar_y_clasificar_directorio(directorio: Path) -> dict:
    """
    Clasifica todos los archivos Excel en un directorio
    
    Args:
        directorio: Ruta del directorio a analizar
        
    Returns:
        Dict con clasificación:
        {
            'ventas': [archivo1, archivo2],
            'inventario': [archivo3],
            'desconocido': [archivo4],
            'no_excel': [archivo5]
        }
    """
    
    resultado = {
        TipoReporte.VENTAS.value.lower(): [],
        TipoReporte.INVENTARIO.value.lower(): [],
        TipoReporte.DESCONOCIDO.value.lower(): [],
        'no_excel': []
    }
    
    directorio = Path(directorio)
    
    if not directorio.exists():
        logger.warning(f"Directorio no existe: {directorio}")
        return resultado
    
    for archivo in directorio.glob('*'):
        if not archivo.is_file():
            continue
        
        if archivo.suffix.lower() not in {'.xlsx', '.xls'}:
            resultado['no_excel'].append(archivo.name)
            continue
        
        try:
            tipo = ReporteDetector.detectar_tipo(archivo)
            key = tipo.value.lower()
            resultado[key].append(archivo.name)
        except Exception as e:
            logger.error(f"Error clasificando {archivo.name}: {str(e)}")
            resultado['desconocido'].append(archivo.name)
    
    return resultado


# ============================================================================
# EJEMPLO DE USO
# ============================================================================

if __name__ == '__main__':
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Probar con un archivo
    test_file = Path('descargas_reportes/ventas/VENTAS_PAGE1_FILA1_20260227_161509.xlsx')
    
    if test_file.exists():
        print(f"\n🔍 Detectando tipo de {test_file.name}...")
        tipo, metadata = ReporteDetector.validar_deteccion(test_file)
        print(f"\n✅ Resultado:")
        print(f"   Tipo: {metadata['tipo']}")
        print(f"   Archivo: {metadata['archivo']}")
        print(f"   Tamaño: {metadata['tamaño_mb']} MB")
        print(f"   Confianza: {metadata['confianza']}")
    else:
        print(f"❌ No se encontró: {test_file}")
